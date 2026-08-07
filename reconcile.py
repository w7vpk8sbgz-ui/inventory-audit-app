"""
재고 대사(Reconciliation) 스크립트
장부재고.xlsx 와 실사재고.xlsx 를 품목코드 기준으로 자동 비교해
차이나는 항목, 누락 항목, 신규(장부에 없는) 항목을 하이라이트한 결과 파일을 생성한다.

사용법:
    python3 reconcile.py
    (필요시 --ledger, --count, --out 옵션으로 파일명 지정 가능)

출력:
    대사결과.xlsx
      - [전체 대사] 시트: 모든 품목 + 차이수량 + 상태
      - [차이항목만] 시트: 차이가 있는 품목만 모아서 바로 확인 가능
"""

import argparse
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # 수량 차이
RED = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")       # 누락(장부에만 있음)
BLUE = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")      # 신규(실사에만 있음)
GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")     # 일치


def classify(row):
    """대사 결과에 따라 상태를 분류한다."""
    if pd.isna(row["장부수량"]):
        return "신규(장부누락)"
    if pd.isna(row["실사수량"]):
        return "실사누락"
    if row["장부수량"] == row["실사수량"]:
        return "일치"
    return "수량차이"


def reconcile(ledger_path, count_path):
    ledger = pd.read_excel(ledger_path)   # 품목코드, 품목명, 장부수량
    count = pd.read_excel(count_path)     # 품목코드, 품목명, 실사수량

    # outer merge: 양쪽에 다 있는 것 + 한쪽에만 있는 것까지 전부 포함
    merged = pd.merge(
        ledger[["품목코드", "품목명", "장부수량"]],
        count[["품목코드", "실사수량"]],
        on="품목코드",
        how="outer",
    )
    # 품목명이 실사 쪽에만 있는 경우(장부에 없는 신규 품목) 보완
    merged["품목명"] = merged["품목명"].fillna(
        merged["품목코드"].map(count.set_index("품목코드")["품목명"])
    )

    merged["차이수량"] = merged["실사수량"] - merged["장부수량"]
    merged["상태"] = merged.apply(classify, axis=1)

    merged = merged.sort_values(by=["상태", "품목코드"]).reset_index(drop=True)
    return merged


def save_with_highlight(df, out_path):
    diff_only = df[df["상태"] != "일치"]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="전체대사", index=False)
        diff_only.to_excel(writer, sheet_name="차이항목만", index=False)

    wb = load_workbook(out_path)
    fill_map = {
        "일치": GREEN,
        "수량차이": YELLOW,
        "실사누락": RED,
        "신규(장부누락)": BLUE,
    }
    status_col_letter = "F"  # 품목코드,품목명,장부수량,실사수량,차이수량,상태 -> F열

    for sheet_name in ["전체대사", "차이항목만"]:
        ws = wb[sheet_name]
        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).font = Font(bold=True)
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=6).value
            fill = fill_map.get(status)
            if fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill
        for col, width in zip("ABCDEF", [12, 22, 12, 12, 10, 16]):
            ws.column_dimensions[col].width = width

    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--ledger", default="장부재고.xlsx")
    parser.add_argument("--count", default="실사재고.xlsx")
    parser.add_argument("--out", default="대사결과.xlsx")
    args = parser.parse_args()

    result = reconcile(args.ledger, args.count)
    save_with_highlight(result, args.out)

    total = len(result)
    diff_count = len(result[result["상태"] != "일치"])
    print(f"총 {total}개 품목 대사 완료. 차이 발견: {diff_count}건")
    print(result[["품목코드", "품목명", "장부수량", "실사수량", "차이수량", "상태"]].to_string(index=False))
    print(f"\n결과 저장: {args.out}")


if __name__ == "__main__":
    main()
